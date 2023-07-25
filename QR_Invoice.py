from tkinter import ttk
import tkinter as tk
import base64
import json
import openpyxl
import tkinter.messagebox as messagebox
from PIL import Image

class Invoice:
    def __init__(self, clave_referencia, ver, fecha, cuit, ptoVta, tipoCmp, nroCmp, importe, moneda, ctz, tipoDocRec, nroDocRec, tipoCodAut, codAut):
        self.clave_referencia = clave_referencia
        self.ver = ver
        self.fecha = fecha
        self.cuit = cuit
        self.ptoVta = ptoVta
        self.tipoCmp = tipoCmp
        self.nroCmp = nroCmp
        self.importe = importe
        self.moneda = moneda
        self.ctz = ctz
        self.tipoDocRec = tipoDocRec
        self.nroDocRec = nroDocRec
        self.tipoCodAut = tipoCodAut
        self.codAut = codAut


def decodificar_codigo(): 
    global invoice

    codigo = caja_texto_escaneado.get()
    lista_lectura = codigo.split('/')
    a_decodificar = lista_lectura[-1][3:]
    mensaje_decodificado = base64.b64decode(a_decodificar).decode('utf-8')
    objeto = json.loads(mensaje_decodificado)
    
     
    clave_referencia = f"{objeto['cuit']}-{objeto['ptoVta']}-{objeto['nroCmp']}-{objeto['tipoCmp']}"
    invoice = Invoice(clave_referencia, objeto['ver'], objeto['fecha'], objeto['cuit'], objeto['ptoVta'], objeto['tipoCmp'],
                      objeto['nroCmp'], objeto['importe'], objeto['moneda'], objeto['ctz'], objeto['tipoDocRec'],
                      objeto['nroDocRec'], objeto['tipoCodAut'], objeto['codAut'])

    etiqueta_cuit["text"] = invoice.cuit
    etiqueta_ptoventa["text"] = invoice.ptoVta
    etiqueta_nrocomp["text"] = invoice.nroCmp
    etiqueta_cod["text"] = invoice.tipoCmp
    etiqueta_subtotal["text"] = "$ " + str(invoice.importe)

    return clave_referencia, invoice

def limpiar_pantalla():
    # Restablecer los valores de las etiquetas a los originales
    etiqueta_cuit["text"] = " "
    etiqueta_ptoventa["text"] = " "
    etiqueta_nrocomp["text"] = " "
    etiqueta_cod["text"] = " "
    etiqueta_subtotal["text"] = " "

    #Vaciar Caja de Texto
    caja_texto_escaneado.delete(0, tk.END)



def registrar_factura():   # OUTPUT, modificable segun usuario final
    global invoice

    # Obtener los valores actuales de las etiquetas y la caja de texto
    cuit = invoice.cuit
    ptoVta = invoice.ptoVta
    nroCmp = invoice.nroCmp
    tipoCmp = invoice.tipoCmp
    importe = invoice.importe
    ver = invoice.ver
    fecha = invoice.fecha
    moneda = invoice.moneda
    cotizacion = invoice.ctz
    tipo_doc_receptor = invoice.tipoDocRec
    nro_doc_receptor = invoice.nroDocRec
    tipo_cod_autorizacion = invoice.tipoCodAut 
    cod_autorizacion = str(invoice.codAut)
    codigo_escaneado = caja_texto_escaneado.get()
    alicuota_iva = int(combo_iva.get())
    monto_subtotal = int(importe) / (1+(alicuota_iva/100))
    monto_iva = int(importe) - monto_subtotal
    campo_clave_factura = str(invoice.cuit) + "-" + str(invoice.ptoVta) + "-" + str(invoice.nroCmp) + "-" + str(invoice.tipoCmp)

    # Crear una instancia de Workbook y cargar el archivo existente o crear uno nuevo si no existe
    try:
        wb = openpyxl.load_workbook("LibroContable.xlsx")
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    # Seleccionar la hoja activa o crear una nueva si no existe
    sheet = wb.active if "Hoja1" in wb.sheetnames else wb.create_sheet()

    # Obtener el número de la última fila
    last_row = sheet.max_row + 1

    # Agregar los datos a una nueva fila en el archivo Excel
    sheet.cell(row=last_row, column=1, value=campo_clave_factura)
    sheet.cell(row=last_row, column=2, value=cuit)
    sheet.cell(row=last_row, column=3, value=ptoVta)
    sheet.cell(row=last_row, column=4, value=nroCmp)
    sheet.cell(row=last_row, column=5, value=tipoCmp)
    sheet.cell(row=last_row, column=6, value=importe)
    sheet.cell(row=last_row, column=7, value=ver)
    sheet.cell(row=last_row, column=8, value=fecha)
    sheet.cell(row=last_row, column=9, value=moneda)
    sheet.cell(row=last_row, column=10, value=cotizacion)
    sheet.cell(row=last_row, column=11, value=tipo_doc_receptor)
    sheet.cell(row=last_row, column=12, value=nro_doc_receptor)
    sheet.cell(row=last_row, column=13, value=tipo_cod_autorizacion)
    sheet.cell(row=last_row, column=14, value=cod_autorizacion)
    sheet.cell(row=last_row, column=15, value=codigo_escaneado)
    sheet.cell(row=last_row, column=16, value=alicuota_iva)
    sheet.cell(row=last_row, column=17, value=monto_subtotal)
    sheet.cell(row=last_row, column=18, value=monto_iva)

    # Guardar los cambios en el archivo Excel
    wb.save("LibroContable.xlsx")

    #Mensaje de Confirmación
    messagebox.showinfo("Factura Registrada", "Factura Registrada!")

    #vaciar datos
    limpiar_pantalla()


##FRONT USER
#Definicion Ventana

ventana = tk.Tk()
ventana.title("Ventana uno")
ventana.geometry('800x600')
ventana.configure(background='black')

logo = tk.PhotoImage(file="ImagenQR.png")
imagen_sub=logo.subsample(3)

fuente_mediana = ("Arial", 10)

#Bienvenidos
etiqueta_bienvenidos = tk.Label(ventana,text="Bienvenidos al Scan QR",font=fuente_mediana, bg="grey", fg="black")
etiqueta_bienvenidos.pack(fill=tk.X, padx= 30, pady= 15)

#Entry
caja_texto_escaneado = tk.Entry(ventana)
caja_texto_escaneado.pack(fill=tk.X, padx= 40, pady= 10)

#Boton Escanear
boton_escanear = tk.Button(ventana, text= "Escanea el código", command= decodificar_codigo) #command= decodificar_codigo(caja_texto_escaneado))
boton_escanear.pack()

boton_limpiar = tk.Button(ventana, text="Limpiar", command=limpiar_pantalla)
boton_limpiar.pack(anchor="center", padx=45, pady=5)

#Logo
etiqueta_logo = ttk.Label(image=imagen_sub)
etiqueta_logo.pack(side = tk.RIGHT, padx= 50, pady= 30)

#ComboBox - IVA
etiqueta_combo_iva = tk.Label(ventana, text= "Tipo de IVA de la factura es %: ", bg="black", fg="white")
etiqueta_combo_iva.pack(anchor="w", padx=30, pady=5)

combo_iva = ttk.Combobox(state="readonly", values=["27", "21", "10,5", "0"])
combo_iva.pack(anchor="w", padx=30, pady=5)


#Etiquetas Datos Factura
etiqueta_cuit_texto = tk.Label(ventana, text= "El CUIT EMISOR ES: ", bg="Black", fg="white")
etiqueta_cuit_texto.pack(anchor="w", padx=30, pady=5)

etiqueta_cuit = tk.Label(ventana)
etiqueta_cuit.pack(anchor="w", padx=35, pady=5)
#---
etiqueta_ptoventa_texto = tk.Label(ventana, text= "El Punto de Venta es: ", bg="black", fg="white")
etiqueta_ptoventa_texto.pack(anchor="w", padx=30, pady=5)

etiqueta_ptoventa = tk.Label(ventana)
etiqueta_ptoventa.pack(anchor="w", padx=35, pady=5)
#----
etiqueta_nrocomp_texto = tk.Label(ventana, text= "El Número de Comprobante es: ", bg="black", fg="white")
etiqueta_nrocomp_texto.pack(anchor="w", padx=30, pady=5)

etiqueta_nrocomp = tk.Label(ventana)
etiqueta_nrocomp.pack(anchor="w", padx=35, pady=5)
#----
etiqueta_cod_texto = tk.Label(ventana, text= "El Código de Comprobante es: ", bg="black", fg="white")
etiqueta_cod_texto.pack(anchor="w", padx=30, pady=5)

etiqueta_cod = tk.Label(ventana)
etiqueta_cod.pack(anchor="w", padx=35, pady=5)
#----
etiqueta_subtotal_texto = tk.Label(ventana, text= "El Subtotal de Comprobante es: ", bg="black", fg="white")
etiqueta_subtotal_texto.pack(anchor="w", padx=30, pady=5)

etiqueta_subtotal = tk.Label(ventana)
etiqueta_subtotal.pack(anchor="w", padx=35, pady=5)

#Boton Registrar
boton_registrar = tk.Button(ventana, text= "REGISTRAR FACTURA", command= registrar_factura) #command= decodificar_codigo(caja_texto_escaneado))
boton_registrar.pack(anchor="center", padx=45, pady=5)

ventana.mainloop()

